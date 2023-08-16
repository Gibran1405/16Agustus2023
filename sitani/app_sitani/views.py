from django.shortcuts import render, redirect
from app_sitani.models import KelompokTani, BeritaPertanian, UsulanBantuanPoktan,TokoPertanian, Profile, BarangDariPemerintah, Golongan, Jabatan
from app_sitani.forms import FormBantuanPoktan,FormProfilePegawai, FormRegister, FormKeltan, FormBerita, FormToko, FormBarangDariPemerintah, LupaSandi, FormGolongan, FormJabatan
from django.contrib import messages 
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.contrib.auth import authenticate,login
from django.contrib.auth.forms import UserCreationForm
from django.db import connection
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.http import HttpResponse,HttpResponseRedirect,HttpRequest,HttpResponseForbidden
from django.http import FileResponse
from django.core.exceptions import ValidationError
import io
from datetime import datetime
from django.contrib.auth.models import User
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from reportlab.lib.pagesizes import landscape, A4, A3   
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT
from reportlab.platypus.tables import Table, TableStyle
from reportlab.platypus import Paragraph,SimpleDocTemplate,Spacer
from django.http import FileResponse
from django.urls import reverse
from django.urls import reverse_lazy
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import datetime
#homepage
def logout_view(request):
    logout(request)
    messages.success(request, 'Anda telah berhasil logout.')
    return redirect('login')
def home(request):
    bertan = BeritaPertanian.objects.all()
    tokper = TokoPertanian.objects.all()
    return render(request, 'index.html', {'bertan': bertan,'tokper': tokper})
def blog(request):
    bertan = BeritaPertanian.objects.all()
    
    for beritapertanian in bertan:
        words = beritapertanian.Isi_Berita.split()
        word_count = len(words)
        beritapertanian.word_count = word_count
    
    context = {
        'bertan': bertan
    }
    return render(request, 'blog.html', context)


def blog_single(request, Id_Berita):
    bertan = BeritaPertanian.objects.filter(Id_Berita=Id_Berita) 
    related_berita = BeritaPertanian.objects.exclude(Id_Berita=Id_Berita)
    return render(request, 'blog-single.html', {'bertan': bertan, 'related_berita': related_berita})


def toko (request):
    tokper = TokoPertanian.objects.all()
    return render(request,'services.html',{'tokper':tokper})
def struktur(request):
    profiles = Profile.objects.all().order_by('Id_Jabatan')  # Mengambil semua profil pegawai dan diurutkan berdasarkan Id_Jabatan
    context = {'profiles': profiles}
    return render(request, 'struktur.html', context)

#createlaporan
@login_required(login_url=settings.LOGIN_URL)
def usulanbantuan_excel(request):
    workbook = Workbook()
    worksheet = workbook.active

    # Define column headers
    headers = ['Nama POKTAN','Nama Bantuan', 'Jumlah Bantuan', 'Satuan Bantuan', 'Tanggal Bantuan', 'Kecamatan', 'Desa', 'Ketua',
               'NIK', 'No Telephone', 'Luas Areal HA', 'Komoditas', 'Status Poktan', 'Kelas Poktan']

    # Write column headers to the worksheet
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    banpoktan = UsulanBantuanPoktan.objects.filter(Status="Diajukan")

    # Write data rows to the worksheet
    for row_num, item in enumerate(banpoktan, 3):
        row = [
            item.Id_Poktan.Nama_Poktan if item.Id_Poktan else '',
            item.Id_Banpem.Nama_Barang if item.Id_Banpem else '',
            item.Jumlah_Bantuan,
            item.Satuan_Bantuan,
            item.Tanggal_Bantuan.strftime('%Y-%m-%d'),
            item.Kecamatan() if item.Id_Poktan else '',
            item.Desa() if item.Id_Poktan else '',
            item.Ketua() if item.Id_Poktan else '',
            item.NIK_Ketua() if item.Id_Poktan else '',
            item.No_Telephone() if item.Id_Poktan else '',
            item.Luas_Areal_HA() if item.Id_Poktan else '',
            ', '.join(item.Komoditas()) if item.Id_Poktan else '',
            item.Status_Poktan() if item.Id_Poktan else '',
            item.Kelas_Poktan() if item.Id_Poktan else '',
        ]

        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value

    # Merge cells for the title
    title_cell = worksheet.cell(row=1, column=1)
    nama_bantuan = UsulanBantuanPoktan.objects.first().Id_Banpem.Nama_Barang.upper() if UsulanBantuanPoktan.objects.exists() else ''
    tahun = datetime.datetime.now().year if UsulanBantuanPoktan.objects.exists() else ''
    ditahun = "DITAHUN {}".format(tahun)
    title_cell.value = "CPCL PENERIMA BANTUAN {} UNTUK KELOMPOK TANI {}".format(nama_bantuan, ditahun)
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')
    worksheet.merge_cells('A1:M1')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="usulan_bantuan.xlsx"'

    workbook.save(response)

    return response

@login_required(login_url=settings.LOGIN_URL)
def bantuanpemerintah_pdf(request):
    buf = io.BytesIO()
    banpem = BarangDariPemerintah.objects.all()

    # Define table data
    data = [['Id Banpem', 'Nama Barang', 'Jumlah Barang', 'Satuan', 'Tanggal']]
    for item in banpem:
        row = [
            item.Id_Banpem,
            item.Nama_Barang,
            item.Jumlah_Barang,
            item.Satuan,
            item.Tanggal,
        ]
        data.append(row)

    # Define table style
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
    ])

    # Create the PDF document with landscape orientation
    pdf = SimpleDocTemplate(buf, pagesize=A4)
    elements = []
     # Add title to the elements list
    title = Paragraph("KELOLA PEMBERITAHUAN BANTUAN UNTUK KELOMPOK TANI", style=ParagraphStyle(name='Title', fontSize=18,spaceAfter=10,alignment=1, leading=18))
    elements.append(title)
    elements.append(Spacer(1, 12))  # Add some vertical space

    # Add table to the elements list
    table = Table(data)
    table.setStyle(style)
    elements.append(table)

    pdf.build(elements)

    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='bantuanpemerintah_pdf.pdf')

@login_required(login_url=settings.LOGIN_URL)
def statusbantuan_pdf(request):
    buf = io.BytesIO()
    banpoktan = UsulanBantuanPoktan.objects.all()

    # Define table data
    data = [['Nama Kelompok Tani','Desa','Kecamatan','Ketua', 'NIK Ketua','No Telephone','Komoditas','Luas Lahan', 'Nama Bantuan','Jumlah Bantuan','Satuan Bantuan','Tanggal Bantuan', 'Status Poktan','Kelas Poktan','Status Pengajuan']]

    for item in banpoktan:
        row = [
            item.Id_Poktan.Nama_Poktan if item.Id_Poktan else None,
            item.Id_Poktan.Desa if item.Id_Poktan else None,
            item.Id_Poktan.Kecamatan if item.Id_Poktan else None,
            item.Id_Poktan.Ketua if item.Id_Poktan else None,
            item.Id_Poktan.NIK_Ketua if item.Id_Poktan else None,
            item.Id_Poktan.No_Telephone if item.Id_Poktan else None,
           ', '.join(item.Id_Poktan.Komoditas) if item.Id_Poktan and item.Id_Poktan.Komoditas else None,        
            item.Id_Poktan.Luas_Areal_HA if item.Id_Poktan else None,
            item.Id_Banpem.Nama_Barang if item.Id_Banpem else None,
            item.Jumlah_Bantuan,
            item.Satuan_Bantuan,
            item.Tanggal_Bantuan,
            item.Status_Poktan() if item.Id_Poktan else None,
            item.Kelas_Poktan() if item.Id_Poktan else None,
            item.Status
        ]
        data.append(row)

    # Define table style
    style = TableStyle([
    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('FONTSIZE', (0, 0), (-1, 0), 4),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
    ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ('LEFTPADDING', (0, 0), (-1, -1), 5),
    ('RIGHTPADDING', (0, 0), (-1, -1), 5),
    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ('FONTSIZE', (0, 0), (-1, -1),8 ),
    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
    ('COLWIDTHS', (0, 0), (-1, -1), [10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]),  # Specify column widths
])


    # Create the PDF document with landscape orientation
    pdf = SimpleDocTemplate(buf, pagesize=landscape(A3))
    elements = []
    # Add title to the elements list
    title = Paragraph("KELOLA USULAN BANTUAN UNTUK KELOMPOK TANI", style=ParagraphStyle(name='Title', fontSize=18,spaceAfter=10,alignment=1, leading=18))
    elements.append(title)
    elements.append(Spacer(1, 12))  # Add some vertical space

    # Add table to the elements list
    table = Table(data)
    table.setStyle(style)
    elements.append(table)

    pdf.build(elements)

    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='Status Bantuan_pdf.pdf')

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Check if the username exists in the database and handle case sensitivity
        try:
            user = User.objects.get(username__iexact=username)
        except User.DoesNotExist:
            messages.error(request, 'Username not found.')
            return redirect('login')
        
        if username.lower() != user.username.lower():
            messages.error(request, 'Username not found.')
            return redirect('login')

        if user.check_password(password):
            login(request, user)
            # Redirect to success page
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid password.')
            # Redirect back to login page with error message
            return redirect('login')

    return render(request, 'registration/login.html')

#forgot password
@login_required(login_url=settings.LOGIN_URL)
def lupasandi(request):
    form = LupaSandi(request.POST)
    return render(request, 'lupasandi.html',{'form':form})

#dashboard
@login_required(login_url=settings.LOGIN_URL)
def dashboard (request):
    kelsitan = KelompokTani.objects.all()
    kelsitan_count = KelompokTani.objects.count()
    context = {
        'kelsitan_count': kelsitan_count,
    }
    tokper = TokoPertanian.objects.all()
    tokper_count = TokoPertanian.objects.count()
    context = {
        'tokper_count': tokper_count,
    }
    bertan = BeritaPertanian.objects.all()
    bertan_count = BeritaPertanian.objects.count()
    context = {
        'bertan_count': bertan_count,
    }
    banpoktan = UsulanBantuanPoktan.objects.all()
    banpoktan_count = UsulanBantuanPoktan.objects.count()
    context = {
        'banpoktan_count': banpoktan_count,
    }
    return render(request, 'dashboard.html', {'kelsitan': kelsitan, 'kelsitan_count': kelsitan_count,'banpoktan_count': banpoktan_count,'bertan_count': bertan_count,'tokper_count': tokper_count})

#kelola kelompok tani
@login_required(login_url=settings.LOGIN_URL)
def keltan(request):
    kelsitan = KelompokTani.objects.all()
    konteks = {
        'kelsitan': kelsitan,
    }
    return render(request, 'keltan.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def tambah_keltan(request):
    if request.POST:
        form = FormKeltan(request.POST, request.FILES)
        if form.is_valid():
    # Lanjutkan dengan logika lainnya
            form.save()
            pesan = "Data berhasil disimpan"
            konteks = {
                'form': form,
                'pesan': pesan,
            }
            return render(request, 'tambah-keltan.html', konteks)
    else:
        form = FormKeltan()

    konteks = {
        'form': form,
    }
    return render(request, 'tambah-keltan.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def ubah_keltan(request, Nama_Poktan):
    kelsitan = KelompokTani.objects.get(Id_Poktan=Nama_Poktan)
    template = 'ubah-keltan.html'

    if request.method == 'POST':
        form = FormKeltan(request.POST or None, request.FILES or None, instance=kelsitan)
        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil diperbaharui")
            return redirect('ubah_keltan', Nama_Poktan=Nama_Poktan)
        else:
            messages.error(request, "Data Gagal diperbaharui")
    else:
        form = FormKeltan(instance=kelsitan)

    konteks = {
        'form': form,
        'kelsitan': kelsitan,
    }
    return render(request, 'ubah-keltan.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def hapus_keltan(request, Nama_Poktan):
    kelsitan = KelompokTani.objects.filter(Id_Poktan=Nama_Poktan)
    kelsitan.delete()
    messages.success(request, "Data berhasil dihapus!")
    return redirect('keltan')


# Kelola User
@login_required(login_url=settings.LOGIN_URL)
def user(request):
    users = User.objects.all()
    konteks = {
        'users' : users
    }
    return render(request,'user.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def tambah_user(request):
    if request.method == 'POST':
        form = FormRegister(request.POST)
        if form.is_valid():
            form.save()
            pesan = "Data berhasil disimpan"
            konteks = {
                'form': form,
                'pesan': pesan,
            }
            return render(request, 'tambah-user.html', konteks)
    else:
        form = FormRegister()

    konteks = {
        'form': form,
    }
    return render(request, 'tambah-user.html', konteks)

#@login_required(login_url=settings.LOGIN_URL)
# def ubah_user(request, id):
#     users = User.objects.get(id=id)
#     template = 'ubah-user.html'
#     if request.method == 'POST':
#         form = FormRegister(request.POST, instance=users)
#         if form.is_valid():
#             form.save()
#             messages.success(request, "Data berhasil diperbaharui")
#             return redirect('ubah_user', id=users.id)

#         else:
#             messages.error(request, "Data Gagal Diperbaharui.")
#     else:
#         form = FormRegister(instance=users)
        
#     konteks = {
#         'form': form,
#         'users': users,
#     }
#     return render(request, template, konteks)

@login_required(login_url=settings.LOGIN_URL)
def hapus_user(request, id):
    users = User.objects.get(id=id)
    users.delete()
    messages.success(request, "Data berhasil dihapus!")
    return redirect('user')


#Kelola Golongan
@login_required(login_url=settings.LOGIN_URL)
def golongan(request):
    golongans = Golongan.objects.all()
    konteks = {
        'golongans' : golongans,
    }
    return render(request, 'golongan.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def tambah_golongan(request):
    if request.POST:
        form = FormGolongan(request.POST)
        if form.is_valid():
    # Lanjutkan dengan logika lainnya
            form.save()
            pesan = "Data berhasil disimpan"
            konteks = {
                'form': form,
                'pesan': pesan,
            }
            return render(request, 'tambah-golongan.html', konteks)
    else:
        form = FormGolongan()

    konteks = {
        'form': form,
    }
    return render(request, 'tambah-golongan.html', konteks)
@login_required(login_url=settings.LOGIN_URL)

def ubah_golongan(request, Id_Golongan):
    golongans = Golongan.objects.get(Id_Golongan=Id_Golongan)
    template = 'ubah-golongan.html'
    if request.method == 'POST':
        form = FormGolongan(request.POST, instance=golongans)
        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil diperbaharui")
            return redirect('ubah_golongan', Id_Golongan=Id_Golongan)
        else:
            messages.error(request, "Data Gagal Diperbaharui.")
    else:
        form = FormGolongan(instance=golongans)
        
    konteks = {
        'form': form,
        'golongans': golongans,
    }
    return render(request, template, konteks)
@login_required(login_url=settings.LOGIN_URL)

def hapus_golongan(request, Id_Golongan):
    golongans = Golongan.objects.get(Id_Golongan=Id_Golongan)
    golongans.delete()
    messages.success(request, "Data berhasil dihapus!")
    return redirect('golongan')


#Kelola Jabatan
@login_required(login_url=settings.LOGIN_URL)
def jabatan(request):
    jabatans = Jabatan.objects.all()
    konteks = {
        'jabatans' : jabatans,
    }
    return render(request, 'jabatan.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def tambah_jabatan(request):
    if request.POST:
        form = FormJabatan(request.POST)
        if form.is_valid():
    # Lanjutkan dengan logika lainnya
            form.save()
            pesan = "Data berhasil disimpan"
            konteks = {
                'form': form,
                'pesan': pesan,
            }
            return render(request, 'tambah-jabatan.html', konteks)
    else:
        form = FormJabatan()

    konteks = {
        'form': form,
    }
    return render(request, 'tambah-jabatan.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def ubah_jabatan(request, Nama_Jabatan):
    jabatans = Jabatan.objects.get(Id_Jabatan=Nama_Jabatan)
    template = 'ubah-jabatan.html'
    if request.method == 'POST':
        form = FormJabatan(request.POST, instance=jabatans)
        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil diperbaharui")
            return redirect('ubah_jabatan', Nama_Jabatan=Nama_Jabatan)
        else:
            messages.error(request, "Data Gagal Diperbaharui.")
    else:
        form = FormJabatan(instance=jabatans)
        
    konteks = {
        'form': form,
        'jabatans': jabatans,
    }
    return render(request, template, konteks) 

@login_required(login_url=settings.LOGIN_URL)
def hapus_jabatan(request,Nama_Jabatan):
    jabatans = Jabatan.objects.get(Id_Jabatan=Nama_Jabatan)
    jabatans.delete()
    messages.success(request, "Data berhasil dihapus!")
    return redirect('jabatan')


#kelola bantuan
@login_required(login_url=settings.LOGIN_URL)
def bantuan(request):
    banpoktan = UsulanBantuanPoktan.objects.filter(Status="Diajukan")
    konteks = {
        'banpoktan': banpoktan,
    }
    return render(request, 'bantuan.html', konteks)

    
@login_required(login_url=settings.LOGIN_URL)
def tambah_bantuan(request):
    if request.method == 'POST':
        form = FormBantuanPoktan(request.POST, request.FILES)
        if form.is_valid():
            bantuan = form.save(commit=False)
            id_poktan = form.cleaned_data.get('Id_Poktan').pk
            jumlah_bantuan = form.cleaned_data.get('Jumlah_Bantuan')

            if id_poktan and jumlah_bantuan:
                try:
                    poktan = KelompokTani.objects.get(pk=id_poktan)
                    barang = BarangDariPemerintah.objects.get(pk=bantuan.Id_Banpem.pk)

                    existing_bantuan_poktan = UsulanBantuanPoktan.objects.filter(
                        Id_Poktan=poktan, Status='Dapat Bantuan'
                    ).exists()

                    if existing_bantuan_poktan:
                        error_message = "Tidak dapat menambahkan Kelompok Tani karena statusnya adalah 'Dapat Bantuan'."
                        messages.error(request, error_message)
                        return redirect('tambah_bantuan')
                    elif barang.Jumlah_Barang > 0 and jumlah_bantuan <= barang.Jumlah_Barang:
                        bantuan.save()
                        barang.Jumlah_Barang -= jumlah_bantuan
                        barang.save()
                        messages.success(request, "Data berhasil disimpan.")
                        return redirect('tambah_bantuan')
                    else:
                        error_message = "Jumlah bantuan tidak valid."
                        messages.error(request, error_message)
                except KelompokTani.DoesNotExist:
                    error_message = "ID Poktan tidak valid."
                    messages.error(request, error_message)
                except BarangDariPemerintah.DoesNotExist:
                    error_message = "Nama barang tidak valid."
                    messages.error(request, error_message)
            else:
                error_message = "Jumlah tidak valid."
                messages.error(request, error_message)
        else:
            messages.error(request, "Data Gagal Diperbaharui")
    else:
        form = FormBantuanPoktan()

    konteks = {
        'form': form,
    }
    return render(request, 'tambah-bantuan.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def ubah_bantuan(request, Id_Bantuan):
    banpoktan = UsulanBantuanPoktan.objects.get(Id_Bantuan=Id_Bantuan)
    template = 'ubah-bantuan.html'
    if request.method == 'POST':
        form = FormBantuanPoktan(request.POST or None, request.FILES or None, instance=banpoktan)
        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil diperbaharui")
            return redirect('ubah_bantuan', Id_Bantuan=Id_Bantuan)
        else:
            messages.error(request, "Data gagal diperbaharui")
    else:
        form = FormBantuanPoktan(instance=banpoktan)

    konteks = {
        'form': form,
        'banpoktan': banpoktan,
    }
    return render(request, template, konteks)   


@login_required(login_url=settings.LOGIN_URL)
def hapus_bantuan(request, Id_Bantuan):
    banpoktan = UsulanBantuanPoktan.objects.get(Id_Bantuan=Id_Bantuan)
    banpoktan.delete()
    messages.success(request, "Data berhasil dihapus!")
    return redirect('bantuan')



#kelola berita
@login_required(login_url=settings.LOGIN_URL)
def beritapertanian(request):
    bertan= BeritaPertanian.objects.all()
    konteks = {
        'bertan': bertan,
    }
    return render(request, 'berita.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def tambah_berita(request):
    if request.POST:
        form = FormBerita(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            form = FormBerita()
            pesan = "Data berhasil disimpan"
            konteks = {
                'form': form,   
                'pesan': pesan,
            }
            return render(request, 'tambah-berita.html', konteks)
    else:
        form = FormBerita()

    konteks = {
        'form': form,
    }
    return render(request, 'tambah-berita.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def ubah_berita(request, Judul_Berita):
    bertan = BeritaPertanian.objects.get(Id_Berita=Judul_Berita)
    template = 'ubah-berita.html'
    if request.method == 'POST':
        form = FormBerita(request.POST or None, request.FILES or None, instance=bertan)
        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil diperbaharui")
            return redirect('ubah_berita', Judul_Berita=Judul_Berita)
        else:   
            messages.success(request, "Data gagal diperbaharui")  
            
    else:
        form = FormBerita(instance=bertan)

    konteks = {
        'form': form,
        'bertan': bertan,
    }
    return render(request, template, konteks)

@login_required(login_url=settings.LOGIN_URL)
def hapus_berita(request, Judul_Berita):
    bertan = BeritaPertanian.objects.filter(Id_Berita=Judul_Berita)
    bertan.delete()
    messages.success(request, "Data berhasil dihapus!")
    return redirect('beritapertanian')


#kelola toko pertanian
@login_required(login_url=settings.LOGIN_URL)
def tokopertanian(request):
    tokper = TokoPertanian.objects.all()
    konteks = {
        'tokper': tokper,
    }
    return render(request, 'tokopertanian.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def tambah_tokopertanian(request):
    if request.POST:
        form = FormToko(request.POST,  request.FILES)
        if form.is_valid():
            form.save()
            form = FormToko()
            pesan = "Data berhasil disimpan"
            konteks = {
                'form': form,
                'pesan': pesan,
            }
            return render(request, 'tambah-toko.html', konteks)

    else:
        form = FormToko()

    konteks = {
        'form': form,
    }
    return render(request, 'tambah-toko.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def ubah_tokopertanian(request, Nama_Toko):
    tokper = TokoPertanian.objects.get(Id_Toko=Nama_Toko)
    template = 'ubah-toko.html'
    if request.POST:
        form = FormToko(request.POST or None, request.FILES or None, instance=tokper)
        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil diperbaharui")
            return redirect('ubah_tokopertanian', Nama_Toko=Nama_Toko)
        else:   
             messages.success(request, "Data gagal diperbaharui")  
    else:
        form = FormToko(instance=tokper)

    konteks = {
        'form': form,
        'tokper': tokper,
    }
    return render(request, template, konteks)

@login_required(login_url=settings.LOGIN_URL)
def hapus_tokopertanian(request, Nama_Toko):
    tokper = TokoPertanian.objects.filter(Id_Toko=Nama_Toko)
    tokper.delete()
    messages.success(request, "Data berhasil dihapus!")
    return redirect('tokopertanian')


#kelola data pegawai
@login_required(login_url=settings.LOGIN_URL)
def datapegawai(request):
    profile = Profile.objects.all()
    konteks = {
        'profile': profile,
    }
    return render(request, 'datapegawai.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def tambah_datapegawai(request):
    if request.method == 'POST':
        form = FormProfilePegawai(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            pesan = "Data berhasil disimpan"
            konteks = {
                'form': form,
                'pesan': pesan,
            }
            return render(request, 'tambah-datapegawai.html', konteks)
    else:
        form = FormProfilePegawai()

    konteks = {
        'form': form,
    }
    return render(request, 'tambah-datapegawai.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def ubah_datapegawai(request, Nama):
    profile = get_object_or_404(Profile, NIP=Nama)
    template = 'ubah-datapegawai.html'
    if request.method == 'POST':
        form = FormProfilePegawai(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil diperbaharui")
            return redirect('ubah_datapegawai', Nama=Nama)
        else:
            messages.error(request, "Data Gagal Diperbaharui.")
    else:
        form = FormProfilePegawai(instance=profile)

    context = {
        'form': form,
        'profile': profile,
    }
    return render(request, template, context)

@login_required(login_url=settings.LOGIN_URL)
def hapus_datapegawai(request, Nama):
    profile = Profile.objects.filter(NIP=Nama)
    profile.delete()
    messages.success(request, "Data berhasil dihapus!")
    return redirect('datapegawai')


#kelola bantuan dari pemerintah
@login_required(login_url=settings.LOGIN_URL)
def banpemtan(request):
    banpem = BarangDariPemerintah.objects.all()
    konteks = {
        'banpem': banpem,
    }
    return render(request, 'banpemtan.html', konteks)

@login_required(login_url=settings.LOGIN_URL)
def tambah_banpemtan(request):
    if request.method == 'POST':
        form = FormBarangDariPemerintah(request.POST)
        if form.is_valid():
    # Lanjutkan dengan logika lainnya
            form.save()
            pesan = "Data berhasil disimpan"
            konteks = {
                'form': form,
                'pesan': pesan,
            }
            return render(request, 'tambah-banpemtan.html', konteks)
    else:
        form = FormBarangDariPemerintah()

    konteks = {
        'form': form,
    }
    return render(request, 'tambah-banpemtan.html', konteks)
@login_required(login_url=settings.LOGIN_URL)
def ubah_banpemtan(request, Nama_Barang):
    banpem = BarangDariPemerintah.objects.get(Id_Banpem=Nama_Barang)
    template = 'ubah-banpemtan.html'
    if request.method == 'POST':
        form = FormBarangDariPemerintah(request.POST, instance=banpem)
        if form.is_valid():
            form.save()
            messages.success(request, "Data berhasil diperbaharui")
            return redirect('ubah_banpemtan', Nama_Barang=Nama_Barang)
        else:
            messages.error(request, "Data Gagal Diperbaharui.")
    else:
        form = FormBarangDariPemerintah(instance=banpem)

    konteks = {
        'form': form,
        'banpem': banpem,
    }
    return render(request, template, konteks)
@login_required(login_url=settings.LOGIN_URL)
def hapus_banpemtan(request, Nama_Barang):
    banpem = BarangDariPemerintah.objects.get(Id_Banpem=Nama_Barang)
    banpem.delete()
    messages.success(request, "Data berhasil dihapus!")
    return redirect('banpemtan')
    

# Penerima Bantuan
@login_required(login_url=settings.LOGIN_URL)
def perban(request):
    banpoktan = UsulanBantuanPoktan.objects.filter(Q(Status='Dapat Bantuan') | Q(Status='Tidak Dapat'))
    konteks = {
        'banpoktan': banpoktan,
    }
    return render(request, 'perban.html', konteks)


